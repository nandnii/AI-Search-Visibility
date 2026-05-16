import os, json, time, re
from datetime import date
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

# ── CONFIG ────────────────────────────────────────────────────────────────────
# Set whichever keys you have. Engines with no key are skipped automatically.
GEMINI_API_KEY     = os.environ.get("GEMINI_API_KEY", "")
OPENAI_API_KEY     = os.environ.get("OPENAI_API_KEY", "")
PERPLEXITY_API_KEY = os.environ.get("PERPLEXITY_API_KEY", "")
ANTHROPIC_API_KEY  = os.environ.get("ANTHROPIC_API_KEY", "")

BRANDS = ["Booking.com", "Airbnb", "Expedia", "Skyscanner", "TripAdvisor", "Hostelworld"]

QUERIES = [
    ("Q01", "best app to book flights online",                   "Transactional"),
    ("Q02", "cheapest way to book hotels last minute",           "Transactional"),
    ("Q03", "best platform to compare flight prices",            "Comparison"),
    ("Q04", "Booking.com vs Airbnb which is better",            "Comparison"),
    ("Q05", "Expedia vs Skyscanner which should I use",         "Comparison"),
    ("Q06", "is Airbnb safe to book through",                   "Trust / Safety"),
    ("Q07", "most trusted travel booking website",              "Trust / Safety"),
    ("Q08", "which travel site has best cancellation policy",   "Trust / Safety"),
    ("Q09", "best app for solo travel planning",                "Use Case"),
    ("Q10", "best platform for booking family holidays",        "Use Case"),
    ("Q11", "best travel app for last minute trips",            "Use Case"),
    ("Q12", "which site has best deals for flights to Europe",  "Use Case"),
    ("Q13", "best alternatives to Booking.com",                 "Discovery"),
    ("Q14", "top travel booking platforms 2025",                "Discovery"),
    ("Q15", "which travel apps do frequent flyers recommend",   "Discovery"),
    ("Q16", "how to find cheapest flights anywhere",            "Planning"),
    ("Q17", "best way to plan a two week Europe trip",          "Planning"),
    ("Q18", "how to book travel without paying extra fees",     "Planning"),
    ("Q19", "best travel rewards credit card companion app",    "Intent / High Value"),
    ("Q20", "which travel platform has best loyalty programme", "Intent / High Value"),
]

# ── ENGINE ADAPTERS ───────────────────────────────────────────────────────────
# Each returns a plain string response. Add/swap models here easily.

def query_gemini(query_text):
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_API_KEY}"
    prompt = f"A user asks: '{query_text}'\n\nGive a helpful, specific answer recommending travel platforms or apps. Be direct and mention specific brand names."
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.8, "maxOutputTokens": 800}
    }
    r = requests.post(url, json=payload, timeout=30)
    r.raise_for_status()
    return r.json()["candidates"][0]["content"]["parts"][0]["text"]

def query_openai(query_text):
    url = "https://api.openai.com/v1/chat/completions"
    headers = {"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"}
    payload = {
        "model": "gpt-4o-mini",
        "messages": [{"role": "user", "content": query_text}],
        "max_tokens": 600
    }
    r = requests.post(url, headers=headers, json=payload, timeout=30)
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"]

def query_perplexity(query_text):
    url = "https://api.perplexity.ai/chat/completions"
    headers = {"Authorization": f"Bearer {PERPLEXITY_API_KEY}", "Content-Type": "application/json"}
    payload = {
        "model": "llama-3.1-sonar-small-128k-online",
        "messages": [{"role": "user", "content": query_text}],
        "max_tokens": 600
    }
    r = requests.post(url, headers=headers, json=payload, timeout=30)
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"]

def query_claude(query_text):
    url = "https://api.anthropic.com/v1/messages"
    headers = {
        "x-api-key": ANTHROPIC_API_KEY,
        "anthropic-version": "2023-06-01",
        "Content-Type": "application/json"
    }
    payload = {
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": 600,
        "messages": [{"role": "user", "content": query_text}]
    }
    r = requests.post(url, headers=headers, json=payload, timeout=30)
    r.raise_for_status()
    return r.json()["content"][0]["text"]

# ── ENGINE REGISTRY ───────────────────────────────────────────────────────────
# Engines are activated automatically if their API key is set.

ALL_ENGINES = [
    ("Gemini",      query_gemini,      GEMINI_API_KEY),
    ("ChatGPT",     query_openai,      OPENAI_API_KEY),
    ("Perplexity",  query_perplexity,  PERPLEXITY_API_KEY),
    ("Claude",      query_claude,      ANTHROPIC_API_KEY),
]

def get_active_engines():
    # Re-read env at runtime in case keys were set after import
    active = []
    key_map = {
        "Gemini":     os.environ.get("GEMINI_API_KEY", ""),
        "ChatGPT":    os.environ.get("OPENAI_API_KEY", ""),
        "Perplexity": os.environ.get("PERPLEXITY_API_KEY", ""),
        "Claude":     os.environ.get("ANTHROPIC_API_KEY", ""),
    }
    fn_map = {
        "Gemini": query_gemini, "ChatGPT": query_openai,
        "Perplexity": query_perplexity, "Claude": query_claude,
    }
    for name, fn in fn_map.items():
        if key_map[name]:
            active.append((name, fn))
    return active

# ── BRAND EXTRACTION ──────────────────────────────────────────────────────────
# Always uses Gemini for extraction (free). Falls back to keyword scan if needed.

def extract_mentions_gemini(response_text):
    if not GEMINI_API_KEY:
        return extract_mentions_fallback(response_text)

    brands_str = ", ".join(BRANDS)
    prompt = f"""You are a data extraction assistant. Analyze this AI response and extract brand visibility data.

Brands to track: {brands_str}

Response to analyze:
\"\"\"{response_text[:1500]}\"\"\"

Return ONLY a valid JSON array. Each element = one brand mention:
- brand: exact brand name from the list
- position: integer (1=first mentioned, 2=second, etc.)
- sentiment: "Positive", "Neutral", or "Negative"
- source_cited: "Yes" or "No"
- snippet: max 15-word phrase from the text mentioning the brand

Only include brands actually mentioned. If none, return [].
No markdown, no explanation. JSON only."""

    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_API_KEY}"
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.1, "maxOutputTokens": 600}
    }
    r = requests.post(url, json=payload, timeout=30)
    r.raise_for_status()
    raw = r.json()["candidates"][0]["content"]["parts"][0]["text"]
    raw = re.sub(r"```json|```", "", raw).strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        match = re.search(r'\[.*\]', raw, re.DOTALL)
        return json.loads(match.group()) if match else extract_mentions_fallback(response_text)

def extract_mentions_fallback(response_text):
    """Simple keyword scan if no Gemini key available."""
    found = []
    text_lower = response_text.lower()
    for i, brand in enumerate(BRANDS):
        if brand.lower() in text_lower:
            pos = text_lower.find(brand.lower())
            found.append({
                "brand": brand, "position": len(found) + 1,
                "sentiment": "Neutral", "source_cited": "No",
                "snippet": response_text[max(0,pos-10):pos+40].strip()
            })
    return sorted(found, key=lambda x: x["position"])

# ── EXCEL WRITER ──────────────────────────────────────────────────────────────

DARK_BG   = "1A1A2E"
ACCENT    = "0F3460"
LIGHT_ROW = "EEF2FF"
WHITE     = "FFFFFF"

def hdr_fill(c): return PatternFill("solid", start_color=c, fgColor=c)
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def sentiment_fill(s):
    if s == "Positive": return PatternFill("solid", start_color="D4EDDA", fgColor="D4EDDA")
    if s == "Negative": return PatternFill("solid", start_color="F8D7DA", fgColor="F8D7DA")
    return PatternFill("solid", start_color="FFF3CD", fgColor="FFF3CD")

def write_excel(rows, active_engine_names, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raw Data"

    ws.merge_cells("A1:K1")
    ws["A1"] = "🔍  AI Search Visibility Tracker — Global Travel Brands"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = hdr_fill(DARK_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:K2")
    engines_str = " · ".join(active_engine_names) or "Manual"
    ws["A2"] = f"Collected: {date.today()}  |  Engines: {engines_str}  |  Brands: {', '.join(BRANDS)}"
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="AAAAAA")
    ws["A2"].fill = hdr_fill("16213E")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    headers = ["query_id","query_text","query_category","engine",
               "brand_mentioned","position","sentiment",
               "source_cited","response_snippet","notes","date_collected"]
    col_widths = [10, 44, 20, 14, 16, 10, 13, 13, 44, 28, 16]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = hdr_fill(ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 22

    for ri, row_data in enumerate(rows, 5):
        sentiment = row_data[6] if len(row_data) > 6 else "Neutral"
        base_fill = hdr_fill(LIGHT_ROW) if ri % 2 == 0 else hdr_fill(WHITE)
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=9)
            c.fill = sentiment_fill(sentiment) if ci == 7 else base_fill
            c.alignment = Alignment(vertical="center", wrap_text=(ci in [2, 9]))
            c.border = thin_border()
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A5"

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.merge_cells("A1:F1")
    ws2["A1"] = "Brand Visibility Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws2["A1"].fill = hdr_fill(DARK_BG)
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 26

    brand_counts   = Counter()
    brand_sent     = {b: {"Positive": 0, "Neutral": 0, "Negative": 0} for b in BRANDS}
    brand_by_engine= {b: Counter() for b in BRANDS}

    for row_data in rows:
        brand, engine, sentiment = row_data[4], row_data[3], row_data[6]
        if brand in BRANDS:
            brand_counts[brand] += 1
            if sentiment in brand_sent[brand]:
                brand_sent[brand][sentiment] += 1
            brand_by_engine[brand][engine] += 1

    summary_headers = ["Brand", "Total Mentions", "Positive", "Neutral", "Negative", "Top Engine"]
    for ci, h in enumerate(summary_headers, 1):
        c = ws2.cell(row=3, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = hdr_fill(ACCENT)
        c.border = thin_border()
        ws2.column_dimensions[get_column_letter(ci)].width = 18

    for ri, brand in enumerate(BRANDS, 4):
        top_engine = brand_by_engine[brand].most_common(1)
        top = top_engine[0][0] if top_engine else "—"
        row_vals = [
            brand,
            brand_counts.get(brand, 0),
            brand_sent[brand]["Positive"],
            brand_sent[brand]["Neutral"],
            brand_sent[brand]["Negative"],
            top
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", bold=(ci==1), size=10)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center")
            if ci == 3: c.fill = PatternFill("solid", start_color="D4EDDA", fgColor="D4EDDA")
            if ci == 4: c.fill = PatternFill("solid", start_color="FFF3CD", fgColor="FFF3CD")
            if ci == 5: c.fill = PatternFill("solid", start_color="F8D7DA", fgColor="F8D7DA")

    wb.save(out_path)
    print(f"\n✅  Saved → {out_path}")
    print(f"📊  Total rows: {len(rows)}  |  Engines used: {', '.join(active_engine_names)}")

# ── MAIN ──────────────────────────────────────────────────────────────────────

def run():
    active_engines = get_active_engines()

    if not active_engines:
        print("❌  No API keys found. Set at least one:")
        print("    export GEMINI_API_KEY='...'")
        print("    export OPENAI_API_KEY='...'")
        print("    export PERPLEXITY_API_KEY='...'")
        print("    export ANTHROPIC_API_KEY='...'")
        return

    print("🚀  AI Search Visibility Agent — Travel Brands")
    print(f"    Active engines : {', '.join(e[0] for e in active_engines)}")
    print(f"    Queries        : {len(QUERIES)}")
    print(f"    Brands tracked : {', '.join(BRANDS)}\n")

    all_rows = []
    today    = str(date.today())

    for i, (q_id, q_text, q_cat) in enumerate(QUERIES, 1):
        for engine_name, engine_fn in active_engines:
            print(f"[{i:02d}/{len(QUERIES)}] [{engine_name}] {q_text[:50]}...")
            try:
                response = engine_fn(q_text)
                time.sleep(50)

                mentions = extract_mentions_gemini(response)
                time.sleep(50)

                if not mentions:
                    print(f"          → no tracked brands mentioned")
                    all_rows.append([q_id, q_text, q_cat, engine_name,
                                     "none", 0, "N/A", "No", "", "", today])
                else:
                    for m in mentions:
                        print(f"          → {m.get('brand','')} (pos:{m.get('position',0)}, {m.get('sentiment','')})")
                        all_rows.append([
                            q_id, q_text, q_cat, engine_name,
                            m.get("brand", ""),
                            m.get("position", 0),
                            m.get("sentiment", "Neutral"),
                            m.get("source_cited", "No"),
                            m.get("snippet", ""),
                            "", today
                        ])

            except Exception as e:
                print(f"          ⚠️  Error: {e}")
                all_rows.append([q_id, q_text, q_cat, engine_name,
                                 "ERROR", 0, "N/A", "No", str(e), "", today])
            time.sleep(50)

    out = "AI_Visibility_Agent_Results.xlsx"
    write_excel(all_rows, [e[0] for e in active_engines], out)

if __name__ == "__main__":
    run()
